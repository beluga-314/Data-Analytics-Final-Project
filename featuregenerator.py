import pandas as pd
from collections import Counter
import re
import pickle
import openpyxl

# Assuming 'your_excel_file.xlsx' is the name of your Excel file
# excel_file = './data_f.xlsx'
excel_file = './data_f copy.xlsx'


with open('stopwords_set.pkl', 'rb') as file:
    stop_words = pickle.load(file)

# Load the Excel file into a DataFrame

# 
def replace_numerical_and_bracket(xlsx_file, sheet_name, column_index):
    workbook = openpyxl.load_workbook(xlsx_file)
    sheet = workbook[sheet_name]
 
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                uppercase_pattern = re.compile(r'[A-Z]')
                cell.value = re.sub(r'(\d+)[\).]|(\&nbsp;)', ',', cell.value)
                cell.value = re.sub(r'[^\w,./+\-:;]', '', cell.value)
                cell.value = re.sub(r',{2,}', ',', cell.value)
                cell.value = re.sub(r'^,+|,+$', '', cell.value)
                cell.value = re.sub(uppercase_pattern, lambda x: x.group().lower(), cell.value)
        
    workbook.save(xlsx_file)

replace_numerical_and_bracket(excel_file, 'Sheet1', 26)
df = pd.read_excel(excel_file)

# Update Position to 'Research Engineer' where Position is NaN and Discipline is 57
df.loc[(pd.isna(df['Position'])) & (df['Discipline'] == 57), 'Position'] = 'Research Engineer'

# Update Position to 'Communication' where Position is NaN and Discipline is 41
df.loc[(pd.isna(df['Position'])) & (df['Discipline'] == 41), 'Position'] = 'Communication'

# Update Position to 'Research Engineer' where Position is NaN and Discipline is 8
df.loc[(pd.isna(df['Position'])) & (df['Discipline'] == 8), 'Position'] = 'Research Engineer'

# Update Position to 'Materials' where Position is NaN and Discipline is 23
df.loc[(pd.isna(df['Position'])) & (df['Discipline'] == 23), 'Position'] = 'Materials'

# Define the mapping of positions
position_mapping = {
 'Member of Technical Staff': 'sde',
 'Communication': 'comm',
 'Program Associate': 'sde',
 'Memory Circuit Design Engineer' : 'vlsi',
 'Semiconductor Technology Modeling Engineer': 'vlsi',
 'Semiconductor Engineer': 'vlsi',
 'Memory Circuit Design Verification Engineer' : 'vlsi',
 'Semiconductor Device Engineer': 'vlsi',
 'Software Engineer 2': 'sde',
 'College Graduate': 'mech',
 'Engineer in Engineering Development Group': 'sde',
 'Data Science FTE - Computational and Data Sciences': 'ds',
 'Design Engineer': 'des',
 'Specialist - Product Development - AI' : 'ds',
 'Graduate Engineer Trainee- Mobile Robotics': 'rob',
 'Graduate Engineer Trainee- Product Design': 'des',
 'Sr. Engineer R&D Mechanical Engineer': 'mech',
 'Engineer, R&D': 'ds',
 'Edison Eng. Development program': 'aero',
 'Data Science associate': 'ds',
 'Data Science Associate Consultant': 'ds',
 'PDF / Trainee': 'chem',
 'Sr. Engineer R&D Power Electronics': 'ee',
 'Jr Research Engineer': 'ds',
 'Physical Design': 'vlsi',
 'Associate Developer': 'sde',
 'Research Associate' : 'sde',
 'VLSI/SoC Design Engineer': 'vlsi',
 'Developer': 'sde',
 'MTS - Modelling Engineer' : 'vlsi',
 'Graduate Engineer': 'mech',
 'Staff Engineer/ Senior Engineer': 'sde',
 'Wireless Baseband Systems': 'comm',
 'Associate - Content Development': 'misc',
 'Senior Associate - Content Development': 'misc',
 'Lead - Experience Design': 'des',
 'Applied Scientist II': 'ds',
 'PTE':'ee',
 'EEDP_Renewables Energy': 'ee',
 'Software Engineer - Machine Learning': 'ds',
 'Sr. Engineer': 'rob',
 'Asst. Professor': 'misc',
 'Jr. Scientist.': 'misc',
 'Software Engineer - Search': 'sde',
 'Intern' : 'misc',
 'Research Scietist': 'ds',
 'Executive - Quant Developer': 'misc',
 'Manager' : 'ds',
 'Machine Learning Engineer': 'ds',
 'Machine Learning Engineer (E2)' : 'ds',
 'Reserach Engineer': 'ds',
 'Computational Modeling & Simulation (CAE)': 'aero',
 'UX Designer/Developer': 'sde',
 'Mechanical Engineering Trainee' : 'mech',
 'Solution Leader (AI Labs)': 'ds',
 'Senior Associate': 'analytics',
 'Manager-Design' : 'des',
 'Mechanical Engineer': 'mech',
 'Robot Learning and Control Researcher' : 'rob',
 'Camera Systems': 'ds',
 'Tech role': 'ds',
 'Program Coordinator': 'misc',
 'hardware': 'vlsi',
 'R & D Scientist': 'ds',
 'Research Scientist': 'ds',
 'QR': 'sde',
 'ML Scientist': 'ds',
 'Ml' : 'ds',
 'Specialist, Product Development & Innovation': 'des',
 'HE' : 'vlsi',
 'H/W' : 'vlsi',
 'MM' : 'mm',
 'Modem Firmware': 'comm',
 'S/W': 'sde',
 'Associate data Scientist': 'ds',
 'PNR': 'vlsi',
 'Firmware': 'comm',
 'RTL': 'vlsi',
 'Digital ' : 'vlsi',
 'ADS' : 'ds',
 'CT AE' : 'ds',
 'Communication System': 'comm',
 'MIEDP' : 'misc',
 'Circuit' : 'vlsi',
 'FLIGHT PHYSICS': 'aero',
 'AIRFRAME (THERMAL)': 'aero',
 'AIRFRAME (NUMERICAL SIMULATION)' : 'aero',
 'Citi Analyt Program': 'analytics',
 'Assitant Manager/ Manager': 'analytics',
 'SME' : 'mech',
 'Computational Engineer': 'mech',
 'Edison Engineering Development Program': 'mech',
 'NLP' : 'ds',
 'Engineering Trainee': 'mech',
 'Lead - XD' : 'des',
 'PE - API': 'chem',
 'PE - Injectables' : 'chem',
 'Analytical Specialist (NMR).' : 'misc',
 'Associate' : 'ds',
 'Associate, Game Design' : 'des',
 'Product Owner': 'des',


 'CRISPR-Cas9 and Translational Biological Research Scientist': 'misc',
 'Asst. Manager - R & D': 'ds',
    'Sr. Engineer - R&D': 'ds',
    'Stress / Design': 'mech',
    'Process Engineering (Injectables)': 'chem',
    'Asst. Product Designer': 'des',
    'Jr. Scientist': 'misc', 
    'Master Teacher': 'misc',
    'Engineer- Hardware': 'vlsi', 
    'DIGITAL': 'vlsi',
    'Product Design Engineer': 'des',
    'ASIC ENGINEER / ARCHITECT': 'vlsi',
    'DIGITAL DESIGN': 'vlsi',
    'Logic Design': 'vlsi',
    'ASIC/VLSI Chip Design': 'vlsi',
    'ANALOG DESIGN': 'vlsi',
    'RFIC Team': 'comm',
    'Circuit Design': 'comm',
    'Modem/ WiFi Systems - Bangalore': 'comm',
    'Modem/ WiFi Systems - Hyderabad': 'comm',
    'MM Systems - Hyderabad': 'mm',
    'MM Systems - Bangalore': 'mm',
    'ML Systems - Bangalore': 'ds',
    'Data Scientist - Mtech': 'ds',
    'Blue Scholar': 'ds',
    'ML Systems - Hyderabad': 'ds',
    'Data Scientist - Ms Research': 'ds',
    'TCAD Engineer': 'mech',
    'Neural Network Engineer': 'ds',
    'MRGR': 'ds',
    'Healthcare': 'ds',
    'PDK': 'vlsi',
    'VLSI': 'vlsi',
    'Data scientist': 'ds',
    'CITI Analyst Program': 'analytics',
    'NUMERICAL SUMILATION': 'mech',
    'Engineering Graduate Programme': 'mech', 
    'GE Renewables': 'ee',
    'Associate Engineer - Airframe': 'aero',
    'Flying Start program for Graduate Engineers and Engineering roles': 'vlsi',
    'BHGE - Data Science & Analytics': 'analytics',
    'BHGE - Industrial': 'mech',
    'BHGE - Digital': 'ds',
    'Software Engineer 3': 'sde',
    'Software Engineer 4': 'sde',
    'Software Engineer 5': 'sde',
    'Software Engineer 6': 'sde',
    'Software Engineer 7': 'sde',
    'Software Engineer 8': 'sde',
    'Software Engineer 9': 'sde',
    'Materials Engineer': 'misc',
    'Margaret Ingels Program': 'misc',
    'Analog Design Engineer': 'vlsi',
    'Computer System Engineer': 'sde',
    'APC Application': 'ee',
    'Lead Data Sciences': 'ds',
    'AI-RESEARCH': 'ds',
    'AIML': 'ds',
    'ENGINEERING': 'ds',
    'ST': 'ds',
    'AIML-ENGINEERING': 'ds',
    'Hardware Engineering': 'vlsi',
    'Research And Innovation': 'aero',
    'Developer Associate': 'sde',
    'BIGDATA DEVELOPMENT ENGINEER': 'sde',
    'Software Engineering': 'sde',
    'S/W Development Engg.': 'sde',
    'SW-AUTO-KERNEL': 'sde',
    'Artificial Intelligence': 'ds',
    'SW-TEGRA-BCR': 'sde',
    'SW-GPU-PUNE': 'sde',
    'Analog - Circuit': 'vlsi',
    'Software Engineer-Trainee': 'sde',
    'PGPT': 'mech',
    'Internship only': 'misc',
    'SE - Research': 'sde',
    'Data Sc & Analytics': 'analytics',
    'Full Stack Developer': 'sde', 
    'API-PE': 'chem',
    'Engineer-Research': 'ds',
    'Design & Engineering - Delhi': 'des',
    'Industerial Design': 'des',
    'Chassis - PGET Engineering': 'mech',
    'Associate Technical Lead-Systems Engineering': 'sde',
    'R&D Engineer': 'ds',
    'Applied Researcher': 'ds',
    'Engineer - Microfluidics': 'mech',
    'Assistant Professor (Senior)': 'misc', 
    'Post Graduate Engineer Trainee': 'mech',
    'Pre-Doctoral Researcher': 'ds',
    'Schlumberger': 'ds',
    'Data not available ': 'da',
    'Controls Engineer': 'ee',
    'Software Engineer Trainee': 'sde',
    'ML Engineer Role': 'ds',
    'Design Engineer I': 'des',
    'Platform Engineer I': 'vlsi',
    'Vantage Research': 'analytics',
    'BXD': 'des',
    'TFG': 'misc',
    'HCL Software': 'sde',
    'Technical Trainee': 'misc',
    'EEDP': 'mech',
    'Associate Data Scientist/ Associate Software Engineer': 'sde',
    'Data Analyst - Data Science': 'sde',
    'SWE- ML 2': 'ds',
    'Assistant Professor': 'misc',
    'CCB Risk Modeling - Applied AI/ML': 'analytics',
    'AI Engineer': 'ds',
    'Data Science': 'ds',
    'UXD': 'sde',
    'PostGraduate Engineer Trainee': 'mech',
    'Associate Engineer': 'ds',
    'ML Scientist Role': 'ds',
    'software development': 'sde',
    'Machine Learning': 'ds',
    'Applied Scientist - 1': 'ds',
    'Machine Learning Engineer ( Speech / NLP)': 'ds',
    'ML': 'ds',
    'SWE': 'sde',
    'Sr Data Scientist': 'ds',
    'CMOR': 'ds',
    'ASIC Engg.': 'vlsi',
    'Software Developer': 'sde',
    'Supervising Associate': 'ds',
    'S/W Engg II': 'sde',
    'ML Systems': 'ds',
    'SDE': 'sde',
    'MTS': 'sde',
    'MM Systems': 'mm',
    'Modem Systems': 'comm',
    'ASIC': 'vlsi',
    'Digital Design': 'vlsi',
    'PD': 'vlsi',
    'HPA': 'comm',
    'Senior Engineer - Analog': 'vlsi',
    'Systems': 'sde',
    'RF/Analog': 'vlsi',
    'CommSys': 'comm',
    'Hardware Engineer (SoC / IP)': 'vlsi',
    'Senior Software Engineer - Embedded': 'rob',
    'System Software Engineer': 'sde',
    'Senior Engineer - S/W': 'sde',
    'Senior Engineer Software - Communications': 'comm',
    'Senior Engineer - Digital': 'vlsi',
    'Senior Electrical Engineer': 'ee',
    'SSD Firmware Engineer': 'vlsi',
    'Senior Mechanical Engineer': 'mech',
    'Software Engineer - Machine Learning': 'ds',
    'Member - Technical Staff': 'sde',
    'AI position': 'ds',
    'SE': 'sde',
    'Mechanical': 'mech',
    'VE- Mechanical': 'mech',
    'Electrical': 'ee',
    'MTS Circuit Design': 'vlsi',
    'Materials': 'misc',
    'AI ML': 'ds',
    'Design Engineer - Analog': 'vlsi',
    'Memory Circuit Design verification Engineer': 'vlsi',
    'Engineer in Engineering Development Group - PhD': 'sde',
    'AI Scientist': 'ds',
    'MTS Logic Design': 'vlsi',
    'S/w Engineer': 'sde',
    'Research': 'aero',
    'Associate Research Engineer': 'ds',
    'Member of Technical Staff I - R&D': 'ds',
    'Data Scientist - 1': 'ds',
    'Data Science Associate': 'ds',
    'Associate Engineer - Aerothermal': 'aero',
    'Associate Engineer - NLFE Stress': 'mech',
    'Research Engineer (AI / Deep Learning - Applied Research)': 'ds',
    'Research Engineer (NxT)': 'ds',
    'Lead - Experince Design': 'des',
    'Product Experience Designer': 'des',
    'Researcher': 'ds',
    'Product Development Engineer I': 'misc',
    'Graduate Trainee Engineer': 'mech',
    'Quantitative Research - Analyst': 'analytics',
    'Back-End Software Engineer': 'sde',
    'Hardware Engineer': 'vlsi',
    'Hardware Engineering Profile': 'vlsi',
    'Machine Learning Systems Engineer -BLR': 'ds',
    'Software Engineer II': 'sde',
    'Multimedia System Engineer -BLR': 'mm',
    'Machine Learning Systems Engineer (Software) -HYD': 'ds',
    'Multimedia System Engineer (Software)-HYD': 'mm',
    'Modem Firmware Engineer-HYD': 'comm',
    'Modem System Engineer- BLR/HYD': 'comm',
    'Member of Technical Staff': 'sde',
    'Software Engineer': 'sde',
    'ML': 'ds',
    'ASIC Engineer': 'vlsi',
    'Hardware': 'vlsi',
    'ASIC HW': 'vlsi',
    'Digital': 'vlsi',
    'Software + Graphic System ': 'sde',
    'Graphics System': 'sde',
    'MTS, Frameworks team': 'sde',
    'Modem System': 'comm',
    'Modem FW': 'comm',
    'Analog': 'vlsi',
    'Embedded Software': 'rob',
    'System Software engineer': 'sde',
    'Applied Scientists I': 'ds',
    'ADI Product': 'ee',
    'digital': 'vlsi',
    'Software Engineer-2': 'sde',
    'Associate Data Scientist': 'ds',
    'Analyst': 'analytics',
    'Senior Data Scientist': 'ds',
    'Technology': 'sde',
    'Analytics': 'analytics',
    'Analog H/W': 'vlsi',
    'Digital H/W': 'vlsi',
    'Communication S/w': 'comm',
    'Staff Engineer': 'sde',
    'MLE': 'ds',
    'MTS2': 'sde',
    'Flight Physics': 'aero',
    'Software Development Engineer 2': 'sde',
    'Data Scientist': 'ds',
    'Sr. Software Engineer': 'sde',
    'Member Research Staff': 'sde',
    'Sr. Physical Design Engineer': 'vlsi',
    'Engineer ': 'mech',
    'Engineer': 'mech',
    'Member Technical Staff': 'sde',
    'Data Engineer': 'ds',
    'Data Scientist.': 'ds',
    'Data Analyst': 'analytics',
    'HW': 'vlsi',
    'SW': 'sde',
    'Engineer in EDG': 'sde',
    'Trainee Engineer': 'mech',
    'Analog Design': 'vlsi',
    'DATA, AI/ML': 'ds',
    'Embedded': 'rob',
    'Computer Science Engineer (CSE)': 'sde',
    'AI/ML': 'ds',
    'PGTE': 'mech',
    'Application Engineer': 'sde',
    'Software Engineer (MTS 1)': 'sde',
    'DS': 'ds',
    'AI': 'ds',
    'Core Researcher ': 'ds',
    'Applied Researcher ': 'ds',
    'Researcher ': 'ds',
    'Engineer 1, Software Design': 'sde',
    'Research Engineer': 'ds',
    'Senior Associate Engineer': 'ds',
    'Lead - Experience Design': 'des',
    'Management Trainee': 'misc',
    'Research Engineer': 'ds',
    'GTE': 'chem',
    'System Engineer': 'sde',
    'Design/process co-optimization and R&D Engineer in advanced nodes ': 'ds',
    'Chip design & advanced design flow engineer': 'ds',
    'PhD Trainee': 'chem', 
    'PGET': 'mech',
    'Lead Engineer': 'ds',
    'Post Graduate engineer trainee': 'ds',
    'AI Engineer': 'ds',
    'AI Engineer': 'ds',
    'Engineer Analog Design': 'vlsi',
    'Analyst 3': 'analytics',
    'Research Engineer ': 'ds',
    'Team Member- New Energy': 'chem',
    'Associate Software engineer': 'sde',
    'Sr. Engineer - Design': 'des',
    'Researcher Trainee': 'ds',
    'Software Development': 'sde',
    'Design': 'des',
    'Python Developer': 'sde',
    'AI & ML': 'ds',
    'R&D': 'ds',
    'Senior': 'ds',
    'Privacy Research Scientist': 'ds',
    'Management/Engineer Trainee': 'misc',
    'PDF Trainee': 'chem',
    'Systems Engineer': 'sde',
    'Data Science Engineer': 'ds',
    'Engineer-I': 'ds',
    'ML Engineer': 'ds',
    'Image Signal Processing Engineer': 'ds',
    'Officer-R & D': 'chem',
    'Product Development Associate.': 'misc',
    'Technical Lead': 'sde',
    'MC Delivery Associate': 'ds',
    'Pre-Doctoral Researcher': 'ds',
    '\xa0AI Engineer': 'ds',
    '\xa0Sr. Engineer - Design': 'des',
    'AI Engineer\xa0': 'ds',
    'CCB Risk Modeling – Applied AI/ML': 'analytics',
    'CCB Risk Modeling – Applied AI/ML': 'analytics',
    'Data Scientist - MTech': 'ds',
    'Data Scientist - MTech': 'ds',
    'Data Scientist - MTech': 'ds',
    'Executive – Quant Developer': 'misc',
    'Lead – Experience Design': 'des',
    'Lead – Experience Design': 'des',
    'ML ': 'ds',
    'Research Engineer\xa0\xa0': 'ds',
    'Research Engineer\xa0\xa0': 'ds',
    'Research Engineer\xa0\xa0': 'ds',
    'Senior Engineer – Analog': 'vlsi',
    'Senior Engineer – Analog': 'vlsi',
    'Senior Engineer – Analog': 'vlsi',
    'Software Engineer – Machine Learning': 'ds',
    'Software Engineer – Machine Learning': 'ds',
    'Software Engineer – Machine Learning': 'ds',
    'Specialist – Product Development - AI': 'ds',
    'Specialist – Product Development - AI': 'ds',
    'SWE-  ML 2': 'ds',
}

numerical_map = {
    'ds': 0,
    'sde': 1,
    'des': 2,
    'ee' : 3,
    'aero': 4,
    'vlsi' : 5,
    'analytics': 6,
    'chem': 7,
    'mech': 8,
    'misc': 9,
    'comm': 10,
    'rob' : 11,
    'mm':12,
    'da':13
}

# Apply the position mapping
df['transformed_position'] = df['Position'].map(position_mapping)

# Initialize an empty dictionary to store courses for each transformed position
transformed_keywords = {}



for transformed_position in df['transformed_position'].unique():

    # Filter DataFrame for the current transformed position
    filtered_df = df[df['transformed_position'] == transformed_position]
    
    # Extract courses from the filtered DataFrame
    courses = filtered_df[['course1', 'course2', 'course3', 'course4', 'course5', 'course6']].values.flatten()

    # Remove NaN values (if any) and concatenate course names
    course_names = ' '.join([str(course) for course in courses if pd.notna(course)])

    # Tokenize course names into words (you can adjust the regex pattern as needed)
    words = re.findall(r'\b\w+\b', course_names.lower())

    # Remove stop words
    words = [word for word in words if word not in stop_words]

    # Count the frequency of each word
    word_counts = Counter(words)

    # Get the most common words up to the specified cutoff
    threshold_cutoff = 30
    most_common_words = word_counts.most_common(n=threshold_cutoff)

    # Extract keywords (words) from the most common words
    keywords = [word for word, count in most_common_words]

    # Store the keywords for the transformed position
    transformed_keywords[transformed_position] = keywords

# skills

skill_keywords = {}
for transformed_position in df['transformed_position'].unique():

    # Filter DataFrame for the current transformed position
    filtered_df = df[df['transformed_position'] == transformed_position]
    
    # Extract courses from the filtered DataFrame
    skills = filtered_df[['Skillset']].values.flatten()

    # Remove NaN values (if any) and concatenate course names
    skill_names = ' '.join([str(skill) for skill in skills if pd.notna(skill)])

    # Tokenize course names into words (you can adjust the regex pattern as needed)
    words = re.findall(r'\b\w+\b', skill_names.lower())

    # Remove stop words
    words = [word for word in words if word not in stop_words]

    # Count the frequency of each word
    word_counts = Counter(words)

    # Get the most common words up to the specified cutoff
    skill_threshold_cutoff = 48
    most_common_words = word_counts.most_common(n=skill_threshold_cutoff)

    # Extract keywords (words) from the most common words
    keywords = [word for word, count in most_common_words]

    # Store the keywords for the transformed position
    skill_keywords[transformed_position + '_skills'] = keywords

    # # Print Transformed Position and Keywords
    # print("Transformed Position:", transformed_position)
    # print("Keywords:", keywords)
    # print()

with open('./course_keywords.pkl', 'wb') as file:
    pickle.dump(transformed_keywords, file)

with open('./skill_keywords.pkl', 'wb') as file:
    pickle.dump(skill_keywords, file)

# Create new columns for each transformed value
transformed_columns = df['transformed_position'].unique()
for transformed_position in transformed_columns:
    df[transformed_position] = 0
    df[transformed_position + '_skills'] = 0

# Iterate through each row and update the new columns based on courses
for index, row in df.iterrows():
    # Extract courses from the current row
    courses = row[['course1', 'course2', 'course3', 'course4', 'course5', 'course6']].values.flatten()

    # Remove NaN values (if any) and concatenate course names
    course_names = ' '.join([str(course) for course in courses if pd.notna(course)])

    # Tokenize course names into words (you can adjust the regex pattern as needed)
    words = re.findall(r'\b\w+\b', course_names.lower())

    # Remove stop words
    words = [word for word in words if word not in stop_words]

    unique_words = set(words)

    # Extract keywords (words) from the unique words
    keywords = list(unique_words)

    # Update the corresponding transformed columns based on the keywords
    for transformed_position in transformed_columns:
        matching_positions = set(keywords) & set(transformed_keywords.get(transformed_position, []))
        
        # Use a cutoff value (you can adjust this value)
        common_cutoff = 5
        if len(matching_positions) >= common_cutoff:
            df.at[index, transformed_position] = 1


# skills
for index, row in df.iterrows():
    # Extract courses from the current row
    skills = row[['Skillset']].values.flatten()

    # Remove NaN values (if any) and concatenate course names
    skill_names = ' '.join([str(skill) for skill in skills if pd.notna(skill)])

    # Tokenize course names into words (you can adjust the regex pattern as needed)
    words = re.findall(r'\b\w+\b', skill_names.lower())

    # Remove stop words
    words = [word for word in words if word not in stop_words]

    unique_words = set(words)

    # Extract keywords (words) from the unique words
    keywords = list(unique_words)

    # Update the corresponding transformed columns based on the keywords
    for transformed_position in transformed_columns:
        matching_positions = set(keywords) & set(skill_keywords.get(transformed_position + '_skills', []))
        
        # Use a cutoff value (you can adjust this value)
        skill_common_cutoff = 5
        if len(matching_positions) >= skill_common_cutoff:
            df.at[index, transformed_position + '_skills'] = 1



columns_to_drop = ['Department', 'Position', 'course1', 'course2', 'course3', 'course4', 'course5', 'course6', 'Skillset' ,'skill_back_end','skill_cloud','skill_database_management','skill_dev_ops','skill_dl','skill_front_end','skill_math','skill_ml','skill_programming','skill_soft_skills','skill_vlsi','skill_web']
# Drop the specified columns
df.drop(columns=columns_to_drop, inplace=True)
df.rename(columns={'transformed_position': 'Sector'}, inplace=True)
# Specify the order of columns
desired_order = ['Slot', 'CTC', 'Sector', *df.columns.difference(['Slot', 'CTC', 'Sector']).tolist()]
# Reorder columns
df = df[desired_order]
df['Sector'] = df['Sector'].map(numerical_map)
df.at[327, 'CTC'] = 1350000
df.at[644, 'CTC'] = 1350000
df.at[1387, 'CTC'] = 1350000
df.at[326, 'CTC'] = 4438200
df.at[351, 'CTC'] = 1200000
df.at[352, 'CTC'] = 1200000
df.at[353, 'CTC'] = 1200000
df.to_csv('final_output.csv', index=False)
